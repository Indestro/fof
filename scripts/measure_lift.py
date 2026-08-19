#!/usr/bin/env python3
"""The lift engine — Next action #1 from intel/t1-edge-for-wing-sourcing.md.

The original engine (5.88% base, 26,636 companies) was run inline and lost; only its JSON
outputs survived, so no later number could be reproduced or compared. This is the committed
replacement. It is deliberately tape-agnostic: point it at any (investor, company, date)
tape and it returns entries/hits/shrunk/lift under one fixed definition.

Standard definition (intel doc §6 — do not vary it silently):
  entries  = seed/pre-seed positions by the investor, inside the window
  hit      = a NEW canonical-tier-1 investor arrives on that company within 24 months
  shrunk   = (hits + PRIOR*base) / (entries + PRIOR), PRIOR = 25
  lift     = shrunk / base           <- always print hits/entries beside it

Adds one dimension the earlier work lacks: EARLINESS. Lift alone still ranks late money —
a crossover that co-invests at Series B scores the same "hit" as an angel who was three
years early. Measured separately (see --earliness), the two are near-orthogonal, so a
preceder must be scored on lift AND earliness, never lift alone.

CAVEAT: absolute lift is NOT comparable across tapes (§8). Only rankings transfer.

Usage:
  python3 measure_lift.py --tape syndicates --min-n 5
  python3 measure_lift.py --tape showcase --min-n 5 --earliness
"""
import argparse, csv, json, math, os, re, warnings
from collections import defaultdict
from datetime import date
warnings.filterwarnings('ignore')

FOF = os.path.expanduser('~/sourcing/fof')
SENSOR = os.path.expanduser('~/sourcing/t1-sensor')
PRIOR = 25.0          # §6 standard prior strength
EARLY_DAYS = 60       # CB announce lag is 1-6mo; <=60d precedence is ambiguous, not causal
SEED_ROUNDS = {'seed', 'pre_seed', 'angel', 'convertible_note', 'Seed', 'Pre-Seed', 'Angel'}


def canonical_t1():
    """Emrah's hand-picked 29 — NOT t1_live_list.json, which omits Coatue/Greylock/CRV/
    Felicis/IVP/Thrive/Norwest/USV/Iconiq. Scoring those as preceders is circular."""
    d = json.load(open(f'{FOF}/intel/t1_canonical_handpicked.json'))
    firms = set(d['firms'])
    aliases = {
        'GV (Google Ventures)': ['GV', 'Google Ventures', 'GOOGLE VENTURES'],
        'Andreessen Horowitz': ['a16z', 'a16z crypto', 'AH Capital'],
        'NEA': ['New Enterprise Associates'],
        'Khosla': ['Khosla Ventures'],
        'First Round': ['First Round Capital'],
        'Felicis Ventures': ['Felicis'],
        'Redpoint Ventures': ['Redpoint'],
        'Tiger Global': ['Tiger Global Management'],
        'Iconiq': ['ICONIQ Growth', 'Iconiq Growth', 'ICONIQ Capital'],
        'Greylock': ['Greylock Partners'],
        'Lightspeed Venture Partners': ['Lightspeed'],
        'Union Square Ventures': ['USV'],
        'Norwest Venture Partners': ['Norwest'],
    }
    for k, vs in aliases.items():
        if k in firms:
            firms.update(vs)
    return firms


def norm(s):
    return re.sub(r'\s+', ' ', str(s or '')).strip()


# ---------------- tapes: each yields (company, investor, round_type, date|None) --------
def tape_syndicates():
    for r in csv.DictReader(open(f'{SENSOR}/syndicates.csv')):
        inv = norm(r['investor'])
        if not inv:
            continue
        try:
            d = date.fromisoformat(r['announced_on'])
        except Exception:
            d = None
        yield r['company_permalink'], inv, r['round_type'], d


def tape_showcase():
    import openpyxl
    WB = os.path.expanduser('~/Downloads/Olympos VC Agent - Showcase Results.xlsx')
    wb = openpyxl.load_workbook(WB, read_only=True)
    seen = set()
    for sn in wb.sheetnames:
        for i, r in enumerate(wb[sn].iter_rows(values_only=True)):
            if i < 3 or not r or len(r) < 29:
                continue
            m = re.match(r'=HYPERLINK\("[^"]*",\s*"?([^")]+)', str(r[1] or ''))
            comp = (m.group(1) if m else str(r[1] or '')).strip()
            if not comp or comp.lower() in seen:
                continue
            seen.add(comp.lower())
            rnd = str(r[8] or '')
            for part in re.split(r'[,;]', str(r[9] or '')):
                inv = norm(part)
                if inv and inv.lower() not in ('none', 'n/a', '-'):
                    yield comp, inv, rnd, None


TAPES = {'syndicates': tape_syndicates, 'showcase': tape_showcase}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--tape', default='showcase', choices=sorted(TAPES))
    ap.add_argument('--min-n', type=int, default=5)
    ap.add_argument('--min-opps', type=int, default=3)
    ap.add_argument('--prior', type=float, default=PRIOR)
    ap.add_argument('--seed-only', action='store_true',
                    help='restrict entries to seed/pre-seed (the §6 standard)')
    ap.add_argument('--earliness', action='store_true')
    ap.add_argument('--top', type=int, default=25)
    ap.add_argument('--out', default=f'{FOF}/backtests/measured_lift.json')
    a = ap.parse_args()

    T1 = canonical_t1()
    T1L = {t.lower() for t in T1}

    by_co = defaultdict(list)
    for co, inv, rnd, d in TAPES[a.tape]():
        by_co[co].append((inv, rnd, d))

    base_hits = 0
    entries, hits = defaultdict(int), defaultdict(int)
    early, opps = defaultdict(set), defaultdict(set)

    for co, recs in by_co.items():
        t1_here = any(i.lower() in T1L for i, _, _ in recs)
        base_hits += t1_here
        dated = [(i, r, d) for i, r, d in recs if d]
        t1_entry = min((d for i, _, d in dated if i.lower() in T1L), default=None)

        seen = set()
        for inv, rnd, d in recs:
            if inv.lower() in T1L or inv in seen:
                continue
            if a.seed_only and rnd not in SEED_ROUNDS:
                continue
            seen.add(inv)
            entries[inv] += 1
            hits[inv] += t1_here
            if t1_entry and d:
                opps[inv].add(co)
                if (t1_entry - d).days > EARLY_DAYS:
                    early[inv].add(co)

    N = len(by_co)
    base = base_hits / N

    out = []
    for inv, n in entries.items():
        if n < a.min_n:
            continue
        k = hits[inv]
        shrunk = (k + a.prior * base) / (n + a.prior)
        e, o = len(early.get(inv, ())), len(opps.get(inv, ()))
        earl = (e + 0.5) / (o + 3.0) if o else None
        out.append(dict(investor=inv, entries=n, hits=k, shrunk=shrunk,
                        lift=shrunk / base, early=e, dated_opps=o, earliness=earl,
                        preceder_score=(shrunk / base) * earl if earl else None))

    out.sort(key=lambda r: -r['lift'])
    json.dump(dict(tape=a.tape, base=base, n_companies=N, prior=a.prior,
                   seed_only=a.seed_only, t1_firms=sorted(T1), investors=out),
              open(a.out, 'w'), indent=1)

    print(f'tape={a.tape}  companies={N}  base P(T1)={base:.2%}  '
          f'investors n>={a.min_n}: {len(out)}/{len(entries)}')
    print(f"\n{'investor':32} {'hits/entries':>13} {'shrunk':>7} {'lift':>6}")
    print('-' * 62)
    for r in out[:a.top]:
        print(f"{r['investor'][:32]:32} {str(r['hits'])+'/'+str(r['entries']):>13} "
              f"{r['shrunk']:7.1%} {r['lift']:6.2f}")

    if a.earliness:
        rk = [r for r in out if r['dated_opps'] >= a.min_opps and r['preceder_score']]
        rk.sort(key=lambda r: -r['preceder_score'])
        print(f"\n=== PRECEDERS: lift x earliness (dated opps >= {a.min_opps}) ===")
        print(f"{'investor':32} {'lift':>6} {'early':>8} {'earl%':>6} {'score':>6}")
        print('-' * 64)
        for r in rk[:a.top]:
            print(f"{r['investor'][:32]:32} {r['lift']:6.2f} "
                  f"{str(r['early'])+'/'+str(r['dated_opps']):>8} "
                  f"{r['earliness']:6.0%} {r['preceder_score']:6.2f}")
        # orthogonality check: does lift predict earliness at all?
        import statistics
        pairs = [(r['lift'], r['earliness']) for r in rk]
        if len(pairs) > 5:
            xs = [p[0] for p in pairs]; ys = [p[1] for p in pairs]
            mx, my = statistics.mean(xs), statistics.mean(ys)
            cov = sum((x-mx)*(y-my) for x, y in pairs)
            den = math.sqrt(sum((x-mx)**2 for x in xs) * sum((y-my)**2 for y in ys))
            print(f"\ncorr(lift, earliness) = {cov/den:+.2f} over {len(pairs)} investors")
    print(f'\nsaved {a.out}')


if __name__ == '__main__':
    main()
